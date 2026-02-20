"""
Excel Spreadsheet Loader.

Excel files are fundamentally different from text documents:
- They contain STRUCTURED data (rows and columns)
- They may have multiple sheets (each is a different "document")
- Cell values can be numbers, dates, formulas, or text

Our strategy: convert each sheet to a readable text format where each row
becomes a natural language sentence. This makes the content searchable
by semantic similarity.

Analogy: Imagine translating a spreadsheet into a paragraph that a human
could read. "In row 3 of the Findings sheet: Region is Hong Kong,
Finding Type is Critical, Deadline is March 31 2026." That text is what
gets embedded and searched.
"""
from langchain_core.documents import Document
import openpyxl
import logging

logger = logging.getLogger(__name__)


def load_xlsx(file_path: str) -> list[Document]:
    """
    Load an Excel file and return LangChain Document objects.

    Each sheet becomes a separate Document. Each row is converted to a
    natural language representation using the header row as field names.

    Args:
        file_path: Path to the .xlsx file

    Returns:
        List of Documents, one per sheet
    """
    documents = []

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows or len(rows) < 2:
                continue  # Skip empty sheets or sheets with only headers

            # First row is headers
            headers = [str(cell or "").strip() for cell in rows[0]]
            sheet_text_parts = [f"Sheet: {sheet_name}\n"]

            for row_idx, row in enumerate(rows[1:], start=2):
                # Convert each row to natural language
                # "Region: Hong Kong | Finding Type: Critical | Deadline: March 31"
                row_parts = []
                for header, cell in zip(headers, row):
                    if cell is not None and str(cell).strip():
                        row_parts.append(f"{header}: {str(cell).strip()}")

                if row_parts:
                    sheet_text_parts.append(" | ".join(row_parts))

            full_text = "\n".join(sheet_text_parts)

            if full_text.strip():
                documents.append(Document(
                    page_content=full_text.strip(),
                    metadata={
                        "source": file_path,
                        "file_type": "xlsx",
                        "sheet_name": sheet_name,
                        "row_count": len(rows) - 1,  # Exclude header
                    }
                ))

        logger.info(f"Loaded XLSX: {file_path} — {len(documents)} sheets")

    except Exception as e:
        logger.error(f"Error loading XLSX {file_path}: {e}")
        raise

    return documents
